import os
from dotenv import load_dotenv
load_dotenv()

import base64
import time
from io import BytesIO

import openpyxl
from PIL import Image
from anthropic import Anthropic

# ── 설정 ──────────────────────────────────────────
DATA_DIR = os.environ["DATA_DIR"]

ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
EXCEL_PATH = os.path.join(DATA_DIR, "유형분류.xlsx")
BASE_DIR   = DATA_DIR

MODEL      = "claude-sonnet-5"
BATCH_SIZE = 5  

TARGET_SUBJECTS = ["과학"]

PILOT_LIMIT = None

client = Anthropic(api_key=ANTHROPIC_API_KEY)


def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def image_to_base64(path: str) -> str:
    img = Image.open(path)
    if img.mode != "RGB":
        img = img.convert("RGB")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


# ── 여러 문제(각자 이미지 포함)를 한 번에 묶어서 해설 생성 ──
def generate_explanation_batch(items, max_retry=3):
    """items: [{row_idx, subject, question, choices, answer, category_large, category_mid, image_full_path}, ...]"""
    tool = {
        "name": "write_explanations",
        "description": "각 문제(row_idx)에 대해, 첨부된 이미지를 참고해서 해설을 작성한다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "row_idx": {"type": "integer"},
                            "explanation": {"type": "string"},
                        },
                        "required": ["row_idx", "explanation"],
                    },
                }
            },
            "required": ["results"],
        },
    }

    # 각 문제마다 [설명 텍스트 + 그 문제의 이미지]를 순서대로 붙여서 content 구성
    content = [{
        "type": "text",
        "text": """아래는 여러 개의 검정고시 문제입니다. 각 문제 설명 바로 다음에 그 문제의 이미지가 첨부되어 있습니다.
각 문제(row_idx)에 대해, 첨부된 이미지를 실제로 참고해서 정답 해설을 작성하세요.

해설 작성 기준:
- 왜 정답이 그 보기인지 핵심 개념을 짚어서 설명 (이미지에 나온 구체적 정보 활용)
- 검정고시 수준(쉬운 설명)에 맞게, 존댓말로, 3~5문장 이내
- 불필요한 서론 없이 바로 설명 시작"""
    }]

    valid_items = []
    for it in items:
        if not os.path.exists(it["image_full_path"]):
            print(f"    [경고] 이미지 파일 없음: {it['image_full_path']} (row {it['row_idx']})")
            continue
        b64 = image_to_base64(it["image_full_path"])
        content.append({
            "type": "text",
            "text": (
                f"\n[row_idx={it['row_idx']}]\n"
                f"과목/유형: {it['subject']} / {it['category_large']} > {it['category_mid']}\n"
                f"문제: {it['question']}\n"
                f"보기: {it['choices']}\n"
                f"정답: {it['answer']}\n"
                f"(아래 이미지가 이 문제의 이미지입니다)"
            ),
        })
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": b64},
        })
        valid_items.append(it)

    content.append({"type": "text", "text": "\n각 문제(row_idx)에 대해 explanation을 작성하세요. 반드시 write_explanations 도구를 호출하세요."})

    if not valid_items:
        return {}

    for attempt in range(max_retry):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=4000,
                tools=[tool],
                tool_choice={"type": "tool", "name": "write_explanations"},
                messages=[{"role": "user", "content": content}],
            )
            for block in resp.content:
                if block.type == "tool_use":
                    data = block.input
                    return {r["row_idx"]: r["explanation"] for r in data["results"]}
        except Exception as e:
            print(f"    [재시도 {attempt + 1}/{max_retry}] {e}")
            time.sleep(8)
    return {}


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # has_image / image_path가 수식일 수 있어 계산된 값을 읽기 위한 참고용 워크북
    wb_values = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_values = wb_values.active

    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["subject", "has_image", "image_path", "question", "choices", "answer",
                "category_large", "category_mid"]
    for r in required:
        if r not in col:
            raise ValueError(f"엑셀에 '{r}' 컬럼이 없습니다. 헤더 확인: {header}")

    if "explanation" not in col:
        new_col_idx = len(header) + 1
        ws.cell(row=1, column=new_col_idx, value="explanation")
        col["explanation"] = new_col_idx - 1

    failed_log = []

    for subject in TARGET_SUBJECTS:
        print(f"\n=== {subject} 처리 시작 ===")

        items = []
        for row_idx_iter, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_subject = row[col["subject"]].value
            if row_subject != subject:
                continue
            image_path = ws_values.cell(row=row_idx_iter, column=col["image_path"] + 1).value
            if not image_path:
                continue

            existing = row[col["explanation"]].value if len(row) > col["explanation"] else None
            if existing:
                continue

            items.append({
                "row_idx": row_idx_iter,
                "subject": subject,
                "question": row[col["question"]].value or "",
                "choices": row[col["choices"]].value or "",
                "answer": row[col["answer"]].value or "",
                "category_large": row[col["category_large"]].value or "",
                "category_mid": row[col["category_mid"]].value or "",
                "image_full_path": os.path.join(BASE_DIR, image_path),
            })

        if PILOT_LIMIT:
            items = items[:PILOT_LIMIT]

        print(f"  처리 대상: {len(items)}건{f' (파일럿 {PILOT_LIMIT}건 제한)' if PILOT_LIMIT else ''}")

        total_batches = (len(items) + BATCH_SIZE - 1) // BATCH_SIZE
        for i, batch in enumerate(chunk(items, BATCH_SIZE), start=1):
            print(f"  배치 {i}/{total_batches} 처리 중...")
            result = generate_explanation_batch(batch)

            got_ids = set(result.keys())
            for it in batch:
                if it["row_idx"] not in got_ids:
                    failed_log.append({
                        "row_idx": it["row_idx"],
                        "subject": subject,
                        "question": it["question"],
                        "사유": "해설 생성 실패 또는 응답 누락",
                    })

            for row_num, explanation in result.items():
                ws.cell(row=row_num, column=col["explanation"] + 1, value=explanation)

            wb.save(EXCEL_PATH)
            time.sleep(1)

        print(f"  {subject} 완료")

    if failed_log:
        fail_wb = openpyxl.Workbook()
        fail_ws = fail_wb.active
        fail_ws.append(["row_idx", "subject", "question", "사유"])
        for item in failed_log:
            fail_ws.append([item["row_idx"], item["subject"], item["question"], item["사유"]])
        fail_ws.column_dimensions["C"].width = 60
        fail_path = EXCEL_PATH.replace(".xlsx", "_해설실패목록.xlsx")
        fail_wb.save(fail_path)
        print(f"\n해설 생성 실패 {len(failed_log)}건 → '{fail_path}'에 저장")
    else:
        print("\n실패 없이 전부 성공")

    print("전체 완료!")


if __name__ == "__main__":
    main()