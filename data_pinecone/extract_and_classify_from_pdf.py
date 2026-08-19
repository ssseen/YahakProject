import os
from dotenv import load_dotenv
load_dotenv()

import re
import json
import base64
import openpyxl
import time
from openai import OpenAI
from pdf2image import convert_from_path
from PIL import Image
from pathlib import Path
from io import BytesIO

# ── 설정 ──────────────────────────────────────────
DATA_DIR = os.environ["DATA_DIR"]
CODE_DIR = os.path.dirname(os.path.abspath(__file__))

OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
BASE_FOLDER    = DATA_DIR
OUTPUT_EXCEL   = os.path.join(DATA_DIR, "유형분류.xlsx")
TAXONOMY_PATH  = os.path.join(CODE_DIR, "taxonomy.md")
ANSWER_FOLDER  = os.path.join(DATA_DIR, "문제모음")
QUESTION_FOLDER = os.path.join(DATA_DIR, "문제모음")

LEVELS = ["고졸", "중졸", "초졸"]
YEARS  = ["2021", "2022", "2023", "2024", "2025", "2026"]

client = OpenAI(api_key=OPENAI_API_KEY)

# ── Taxonomy 로드 및 파싱 ──────────────────────────
with open(TAXONOMY_PATH, "r", encoding="utf-8") as f:
    TAXONOMY_RAW = f.read()

def parse_taxonomy(taxonomy_text: str) -> dict:
    """taxonomy.md에서 과목별 대분류/중분류 목록 추출"""
    result = {}
    current_section = None

    for line in taxonomy_text.split("\n"):
        # ## 초등_수학 같은 섹션 헤더
        section_match = re.match(r"^## (.+)", line)
        if section_match:
            current_section = section_match.group(1).strip()
            result[current_section] = {"large": set(), "mid": set()}
            continue

        # - 대분류 > 중분류 형식
        if current_section and line.strip().startswith("- "):
            parts = line.strip()[2:].split(" > ")
            if len(parts) >= 2:
                large = parts[0].strip()
                mid = parts[1].strip()
                result[current_section]["large"].add(large)
                result[current_section]["mid"].add(mid)
            elif len(parts) == 1:
                large = parts[0].strip()
                result[current_section]["large"].add(large)

    # set → list 변환
    for section in result:
        result[section]["large"] = sorted(list(result[section]["large"]))
        result[section]["mid"] = sorted(list(result[section]["mid"]))

    return result

TAXONOMY = parse_taxonomy(TAXONOMY_RAW)

def get_taxonomy_for(exam_type: str, subject: str) -> dict:
    """과목에 맞는 taxonomy 반환"""
    level_map = {"초졸": "초등", "중졸": "중등", "고졸": "고등"}
    level = level_map.get(exam_type, "")

    # 영어는 초중고 동일
    if subject == "영어":
        key = "초등_영어 / 중등_영어 / 고등_영어 (동일 구조)"
        if key in TAXONOMY:
            return TAXONOMY[key]

    key = f"{level}_{subject}"
    return TAXONOMY.get(key, {"large": [], "mid": []})

# ── 엑셀 초기화 ────────────────────────────────────
if Path(OUTPUT_EXCEL).exists():
    wb = openpyxl.load_workbook(OUTPUT_EXCEL)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "id", "exam_type", "year", "exam_round", "subject",
        "question_number", "question", "choices", "answer",
        "has_image", "image_path",
        "category_large", "category_mid", "category_small"
    ])

# ── 이미 처리된 id 목록 ────────────────────────────
processed_ids = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        processed_ids.add(row[0])

# ── 페이지 위치 계산 (합본 PDF 구조) ───────────────
EXAMS_PER_LEVEL = 11

def get_page_count(subject: str, exam_type: str) -> int:
    """이 과목/급수 조합의 문제 PDF 페이지 수 (국어 고졸·중졸만 예외로 4페이지)"""
    if subject == "국어" and exam_type in ("고졸", "중졸"):
        return 4
    return 2

def get_start_page_index(subject: str, exam_type: str, year: str, exam_round: str) -> int:
    """{subject}전체.pdf 안에서 해당 시험이 시작하는 페이지 인덱스(0부터 시작)"""
    if exam_type not in LEVELS:
        raise ValueError(f"알 수 없는 급수: {exam_type}")

    page_count = get_page_count(subject, exam_type)

    # 앞선 급수들이 차지한 총 페이지 수
    offset = 0
    for lvl in LEVELS:
        if lvl == exam_type:
            break
        offset += get_page_count(subject, lvl) * EXAMS_PER_LEVEL

    # 같은 급수 안에서 연도/회차 위치
    year_idx = YEARS.index(str(year))
    if year_idx < 5:  # 2021~2025: 회차 1,2
        within_offset = (year_idx * 2 + (int(exam_round) - 1)) * page_count
    else:
        within_offset = (5 * 2) * page_count

    return offset + within_offset

def get_answer_page_index(exam_type: str, year: str, exam_round: str) -> int:
    """{subject}_전체답안.pdf 안에서 해당 시험 답이 있는 페이지 인덱스(0부터 시작, 답안은 시험당 1페이지 고정)"""
    if exam_type not in LEVELS:
        raise ValueError(f"알 수 없는 급수: {exam_type}")

    level_offset = LEVELS.index(exam_type) * EXAMS_PER_LEVEL

    year_idx = YEARS.index(str(year))
    if year_idx < 5:
        within_offset = year_idx * 2 + (int(exam_round) - 1)
    else:
        within_offset = 5 * 2

    return level_offset + within_offset

# ── 이미지 → base64 변환 ──────────────────────────
def image_to_base64(img: Image.Image) -> str:
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")

# ── GPT: 페이지 분류 (재시도 포함) ───────────────
def classify_page(page_image: Image.Image, info: dict, page_num: int, max_retry=3) -> list:
    tax = get_taxonomy_for(info["exam_type"], info["subject"])
    large_list = tax["large"]
    mid_list = tax["mid"]

    prompt = f"""아래는 검정고시 {info['exam_type']} {info['subject']} 시험지 {page_num}페이지 이미지입니다.

이 페이지에 있는 각 문제를 분석하여 아래 JSON 배열로만 답하세요.
다른 텍스트 없이 JSON만 출력하세요.

사용 가능한 대분류 목록 (이 중에서만 선택):
{large_list}

사용 가능한 중분류 목록 (이 중에서만 선택):
{mid_list}

주의사항:
- category_large는 반드시 위 대분류 목록 중 하나를 그대로 사용
- category_mid는 반드시 위 중분류 목록 중 하나를 그대로 사용
- 목록에 없는 단어 절대 사용 금지. 가장 유사한 항목으로 대체
- category_small은 문제 핵심 개념 2~4글자 자유롭게 작성 (없으면 -)
- has_image는 문제에 도형/표/그래프 등 이미지가 있으면 true, 없으면 false
- 보기가 이미지(도형/그래프 등)인 경우 choices는 빈 문자열로 표시
- category_small은 taxonomy에 소분류가 있으면 반드시 taxonomy의 소분류에서 선택, 없으면 문제 핵심 개념 2~4글자 자유롭게 작성

출력 형식:
[
  {{
    "question_number": 1,
    "question": "문제 텍스트 (보기 제외)",
    "choices": "①...|②...|③...|④...",
    "has_image": true,
    "category_large": "대분류",
    "category_mid": "중분류",
    "category_small": "소분류 또는 -"
  }}
]"""

    b64 = image_to_base64(page_image)

    for attempt in range(max_retry):
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}}
                        ]
                    }
                ],
                max_tokens=4000
            )
            text = response.choices[0].message.content.strip()
            text = re.sub(r"```json|```", "", text).strip()
            match = re.search(r'\[.*\]', text, re.DOTALL)
            if match:
                text = match.group(0)
            return json.loads(text)
        except Exception as e:
            print(f"  [재시도 {attempt+1}/{max_retry}] {e}")
            time.sleep(10)

    return []

# ── 답안지 파싱 (합본 답안 PDF에서 해당 페이지만 읽기) ──
answer_pdf_cache = {}

def parse_answer(exam_type: str, year: str, exam_round: str, subject: str) -> dict:
    if subject not in answer_pdf_cache:
        answer_path = os.path.join(ANSWER_FOLDER, f"{subject}_전체답안.pdf")
        if not os.path.exists(answer_path):
            print(f"  [답안파일없음] {answer_path}")
            answer_pdf_cache[subject] = None
        else:
            answer_pdf_cache[subject] = convert_from_path(answer_path)

    pages = answer_pdf_cache[subject]
    if pages is None:
        return {}

    page_idx = get_answer_page_index(exam_type, year, exam_round)
    if page_idx >= len(pages):
        print(f"  [답안페이지범위초과] {subject} idx={page_idx}, 전체 {len(pages)}페이지")
        return {}

    page_img = pages[page_idx]
    b64 = image_to_base64(page_img)
    prompt = f"""이 답안지 이미지에서 {subject} 과목의 문제번호와 정답만 추출하세요.
다른 텍스트 없이 JSON으로만 답하세요.
형식: {{"1": "3", "2": "1", "3": "4", ...}}
정답은 숫자로만 표시하세요."""

    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}}
                        ]
                    }
                ],
                max_tokens=500
            )
            text = response.choices[0].message.content.strip()
            text = re.sub(r"```json|```", "", text).strip()
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if match:
                text = match.group(0)
            answers = json.loads(text)
            time.sleep(3)
            return answers
        except Exception as e:
            print(f"  [답안재시도] {e}")
            time.sleep(10)

    return {}

# ── 메인 처리 ──────────────────────────────────────
TEST_MODE = False  # True면 첫 조합 1개만 처리, False면 전체 처리
SUBJECTS = ["국어", "수학", "사회", "과학", "영어"]

question_pdf_cache = {}

for subject in SUBJECTS:
    pdf_path = os.path.join(QUESTION_FOLDER, f"{subject}전체.pdf")
    if not os.path.exists(pdf_path):
        print(f"[파일없음] {pdf_path}")
        continue

    print(f"\n=== {subject} 처리 시작 ===")
    all_pages = convert_from_path(pdf_path)

    for exam_type in LEVELS:
        for year in YEARS:
            rounds_this_year = ["1"] if year == "2026" else ["1", "2"]

            for exam_round in rounds_this_year:
                if TEST_MODE and not (subject == "국어" and exam_type == "고졸" and year == "2021" and exam_round == "1"):
                    continue

                info = {
                    "exam_type": exam_type,
                    "year": year,
                    "exam_round": exam_round,
                    "subject": subject,
                }

                check_id = f"{exam_type}_{year}_{exam_round}회_{subject}_1"
                if check_id in processed_ids:
                    print(f"[이미처리됨] {exam_type} {year} {exam_round}차 {subject}")
                    continue

                start_idx = get_start_page_index(subject, exam_type, year, exam_round)
                page_count = get_page_count(subject, exam_type)
                target_pages = all_pages[start_idx: start_idx + page_count]

                if not target_pages:
                    print(f"  [페이지없음] {exam_type} {year} {exam_round}차 {subject} (idx={start_idx})")
                    continue

                print(f"\n처리중: {exam_type} {year} {exam_round}차 {subject} (페이지 {start_idx+1}~{start_idx+page_count})")

                answers = parse_answer(exam_type, year, exam_round, subject)
                print(f"  답안 로드: {len(answers)}개")

                for page_num, page_img in enumerate(target_pages, start=1):
                    print(f"  페이지 {page_num}/{len(target_pages)}")

                    results = classify_page(page_img, info, page_num)
                    if not results:
                        print(f"  [스킵] 페이지 {page_num} 분류 실패")
                        continue

                    time.sleep(5)

                    for r in results:
                        q_num   = r.get("question_number", 0)
                        has_img = r.get("has_image", False)
                        answer  = answers.get(str(q_num), "")
                        row_id  = f"{exam_type}_{year}_{exam_round}회_{subject}_{q_num}"

                        ws.append([
                            row_id,
                            exam_type,
                            year,
                            exam_round,
                            subject,
                            q_num,
                            r.get("question", ""),
                            r.get("choices", ""),
                            answer,
                            has_img,
                            "",
                            r.get("category_large", ""),
                            r.get("category_mid", ""),
                            r.get("category_small", ""),
                        ])

                wb.save(OUTPUT_EXCEL)
                print(f"  저장완료: {exam_type} {year} {exam_round}차 {subject}")

print("\n전체 완료!")