import os
from dotenv import load_dotenv
load_dotenv()

import re
import openpyxl

DATA_DIR = os.environ["DATA_DIR"]
CODE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_PATH    = os.path.join(DATA_DIR, "유형분류.xlsx")
TAXONOMY_PATH = os.path.join(CODE_DIR, "taxonomy.md")
OUTPUT_PATH   = os.path.join(DATA_DIR, "구조검증_오류행.xlsx")


# ── Taxonomy 파싱 (대분류 -> 중분류 리스트, 쌍 유지) ─────
def parse_taxonomy(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    result = {}
    section = None

    for line in text.split("\n"):
        m = re.match(r"^## (.+)", line)
        if m:
            section = m.group(1).strip()
            result[section] = {}
            continue

        if section and line.strip().startswith("- "):
            parts = line.strip()[2:].split(" > ")
            if len(parts) >= 2:
                large, mid = parts[0].strip(), parts[1].strip()
                result[section].setdefault(large, [])
                if mid not in result[section][large]:
                    result[section][large].append(mid)
            elif parts and parts[0].strip():
                large = parts[0].strip()
                result[section].setdefault(large, [])

    return result


TAXONOMY = parse_taxonomy(TAXONOMY_PATH)


def resolve_taxonomy(exam_type, subject):
    if subject == "영어":
        for key in TAXONOMY:
            if "영어" in key:
                return TAXONOMY[key]
        return None
    level_map = {"초졸": "초등", "중졸": "중등", "고졸": "고등"}
    level = level_map.get(exam_type, "")
    return TAXONOMY.get(f"{level}_{subject}")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["exam_type", "subject", "question", "choices",
                "category_large", "category_mid"]
    for r in required:
        if r not in col:
            raise ValueError(f"엑셀에 '{r}' 컬럼이 없습니다. 헤더 확인: {header}")

    error_rows = []

    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        exam_type = row[col["exam_type"]].value
        subject = row[col["subject"]].value
        question = row[col["question"]].value or ""
        choices = row[col["choices"]].value or ""
        large = row[col["category_large"]].value
        mid = row[col["category_mid"]].value

        if not exam_type or not subject:
            continue

        if not large or not mid:
            error_rows.append((row_num, exam_type, subject, question, choices,
                                large or "", mid or "", "대분류/중분류 값이 비어있음"))
            continue

        tax = resolve_taxonomy(exam_type, subject)
        if tax is None:
            error_rows.append((row_num, exam_type, subject, question, choices,
                                large, mid, "taxonomy 섹션 없음"))
            continue

        if large not in tax:
            error_rows.append((row_num, exam_type, subject, question, choices,
                                large, mid, "대분류 자체가 목록에 없음"))
            continue

        if mid not in tax[large]:
            error_rows.append((row_num, exam_type, subject, question, choices,
                                large, mid, f"중분류가 '{large}' 하위에 속하지 않음"))
            continue

    # ── 결과를 별도 엑셀로 저장 ──
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["row_num", "exam_type", "subject", "question", "choices",
                   "category_large", "category_mid", "오류 사유"])

    for r in error_rows:
        out_ws.append(list(r))

    out_ws.column_dimensions["D"].width = 60  # question
    out_ws.column_dimensions["E"].width = 40  # choices
    out_ws.column_dimensions["H"].width = 35  # 오류 사유

    out_wb.save(OUTPUT_PATH)

    print(f"구조 검증 완료: 총 {len(error_rows)}건의 오류 발견")
    print(f"→ '{OUTPUT_PATH}'에 저장했습니다.")
    if len(error_rows) == 0:
        print("(0건이면 정상 — 2단계 enum 강제 방식이 의도대로 작동한 것)")


if __name__ == "__main__":
    main()