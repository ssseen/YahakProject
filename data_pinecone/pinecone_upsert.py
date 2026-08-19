import os
from dotenv import load_dotenv
load_dotenv()

import hashlib

import openpyxl
from pinecone import Pinecone

# ── 설정 ──────────────────────────────────────────
DATA_DIR = os.environ["DATA_DIR"]

PINECONE_API_KEY = os.environ["PINECONE_API_KEY"]
EXCEL_PATH = os.path.join(DATA_DIR, "유형분류.xlsx")

INDEX_NAME       = "geondi-questions"   
NAMESPACE        = "questions"
BATCH_SIZE = 90

pc = Pinecone(api_key=PINECONE_API_KEY)


# ── 1. 인덱스 생성 ─────────────
def ensure_index():
    existing = [idx["name"] for idx in pc.list_indexes()]
    if INDEX_NAME in existing:
        print(f"인덱스 '{INDEX_NAME}' 이미 존재 — 재사용")
        return

    print(f"인덱스 '{INDEX_NAME}' 생성 중...")
    pc.create_index_for_model(
        name=INDEX_NAME,
        cloud="aws",
        region="us-east-1",
        embed={
            "model": "llama-text-embed-v2", 
            "field_map": {"text": "text"}, 
        },
    )
    print("인덱스 생성 완료")


# ── 2. 엑셀 → 레코드 리스트 변환 ─────────────────
def build_records():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["id", "exam_type", "subject", "year", "exam_round",
                "question_number", "question", "choices", "answer",
                "category_large", "category_mid"]
    for r in required:
        if r not in col:
            raise ValueError(f"엑셀에 '{r}' 컬럼이 없습니다. 헤더 확인: {header}")

    has_small = "category_small" in col
    has_explanation = "explanation" in col

    records = []
    for row in ws.iter_rows(min_row=2):
        row_id = row[col["id"]].value
        question = row[col["question"]].value or ""
        choices = row[col["choices"]].value or ""

        if not row_id or not question:
            continue

        # 임베딩 대상 텍스트: 문제 + 보기
        text = f"{question}\n{choices}".strip()

        ascii_id = hashlib.md5(str(row_id).encode("utf-8")).hexdigest()

        metadata = {
            "original_id": str(row_id),
            "exam_type": row[col["exam_type"]].value or "",
            "subject": row[col["subject"]].value or "",
            "year": str(row[col["year"]].value or ""),
            "exam_round": str(row[col["exam_round"]].value or ""),
            "question_number": str(row[col["question_number"]].value or ""),
            "answer": str(row[col["answer"]].value or ""),
            "category_large": row[col["category_large"]].value or "",
            "category_mid": row[col["category_mid"]].value or "",
        }
        if has_small:
            metadata["category_small"] = row[col["category_small"]].value or ""
        if has_explanation:
            metadata["explanation"] = row[col["explanation"]].value or ""

        records.append({
            "id": ascii_id,
            "text": text,
            **metadata,
        })

    return records


# ── 3. 배치 업서트 ────────────────────────────────
def upsert_records(records):
    index = pc.Index(INDEX_NAME)

    total = len(records)
    for i in range(0, total, BATCH_SIZE):
        batch = records[i:i + BATCH_SIZE]
        index.upsert_records(namespace=NAMESPACE, records=batch)
        print(f"  업서트 진행: {min(i + BATCH_SIZE, total)}/{total}")

    print(f"\n전체 {total}건 업서트 완료")


def main():
    ensure_index()
    records = build_records()
    print(f"엑셀에서 {len(records)}건 로드")
    upsert_records(records)


if __name__ == "__main__":
    main()