import os
from dotenv import load_dotenv
load_dotenv()

import time
import openpyxl
from anthropic import Anthropic

# ── 설정 ──────────────────────────────────────────
DATA_DIR = os.environ["DATA_DIR"]

ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
EXCEL_PATH = os.path.join(DATA_DIR, "유형분류.xlsx")

MODEL      = "claude-sonnet-5"
BATCH_SIZE = 10

# 이번에 처리할 과목만 지정. 예: ["수학"] 또는 ["수학", "영어"]. 전체는 None
TARGET_SUBJECTS = None

TEXT_ONLY_SUBJECTS = {"사회"}

# 파일럿 테스트용: 몇 건만 먼저 돌려보고 품질 확인할 때 숫자 지정 (예: 15). 전체 실행은 None
PILOT_LIMIT = None

client = Anthropic(api_key=ANTHROPIC_API_KEY)


def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def generate_explanation_batch(items, max_retry=3):
    tool = {
        "name": "write_explanations",
        "description": "각 문제(row_idx)에 대한 해설을 작성한다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "row_idx": {"type": "integer"},
                            "explanation": {"type": "string"},
                        },
                        "required": ["row_idx", "explanation"],
                    },
                }
            },
            "required": ["results"],
        },
    }

    items_text = "\n\n".join(
        f"[row_idx={it['row_idx']}]\n"
        f"과목/유형: {it['subject']} / {it['category_large']} > {it['category_mid']}\n"
        f"문제: {it['question']}\n"
        f"보기: {it['choices']}\n"
        f"정답: {it['answer']}"
        for it in items
    )

    prompt = f"""아래는 검정고시 문제들입니다. 각 문제(row_idx)에 대해 정답 해설을 작성하세요.

해설 작성 기준:
- 왜 정답이 그 보기인지 핵심 개념을 짚어서 설명
- 검정고시 수준(쉬운 설명)에 맞게, 존댓말로, 3~5문장 이내
- 오답 보기에 대한 설명은 꼭 필요한 경우만 짧게 언급
- 불필요한 서론 없이 바로 설명 시작

문제 목록:
{items_text}

각 문제(row_idx)에 대해 explanation을 작성하세요. 반드시 write_explanations 도구를 호출하세요."""

    for attempt in range(max_retry):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=4000,
                tools=[tool],
                tool_choice={"type": "tool", "name": "write_explanations"},
                messages=[{"role": "user", "content": prompt}],
            )
            for block in resp.content:
                if block.type == "tool_use":
                    data = block.input
                    return {r["row_idx"]: r["explanation"] for r in data["results"]}
        except Exception as e:
            print(f"  [재시도 {attempt + 1}/{max_retry}] {e}")
            time.sleep(8)
    return {}


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # has_image가 수식으로 저장된 경우 대비: 계산된 값을 읽기 위한 참고용 워크북
    wb_values = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_values = wb_values.active

    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["subject", "has_image", "question", "choices", "answer",
                "category_large", "category_mid"]
    for r in required:
        if r not in col:
            raise ValueError(f"엑셀에 '{r}' 컬럼이 없습니다. 헤더 확인: {header}")

    if "explanation" not in col:
        new_col_idx = len(header) + 1
        ws.cell(row=1, column=new_col_idx, value="explanation")
        col["explanation"] = new_col_idx - 1

    items = []
    for row_idx_iter, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_num = row[0].row
        subject = row[col["subject"]].value
        has_image = ws_values.cell(row=row_idx_iter, column=col["has_image"] + 1).value

        if TARGET_SUBJECTS and subject not in TARGET_SUBJECTS:
            continue
        
        if subject not in TEXT_ONLY_SUBJECTS and has_image in (True, "True", "TRUE", "true", 1):
            continue

        existing = row[col["explanation"]].value if len(row) > col["explanation"] else None
        if existing:
            continue

        items.append({
            "row_idx": row_num,
            "subject": subject or "",
            "question": (row[col["question"]].value or "")[:600],
            "choices": row[col["choices"]].value or "",
            "answer": row[col["answer"]].value or "",
            "category_large": row[col["category_large"]].value or "",
            "category_mid": row[col["category_mid"]].value or "",
        })

    if PILOT_LIMIT:
        items = items[:PILOT_LIMIT]

    print(f"[텍스트 해설] 대상 과목: {TARGET_SUBJECTS or '전체'} / 처리 대상: {len(items)}건"
          f"{f' (파일럿 {PILOT_LIMIT}건 제한)' if PILOT_LIMIT else ''}")

    total_batches = (len(items) + BATCH_SIZE - 1) // BATCH_SIZE
    for i, batch in enumerate(chunk(items, BATCH_SIZE), start=1):
        print(f"  배치 {i}/{total_batches} 처리 중...")
        result = generate_explanation_batch(batch)

        for row_num, explanation in result.items():
            ws.cell(row=row_num, column=col["explanation"] + 1, value=explanation)

        wb.save(EXCEL_PATH)
        time.sleep(1)

    print("\n완료!")


if __name__ == "__main__":
    main()