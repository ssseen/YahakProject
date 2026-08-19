import os
from dotenv import load_dotenv
load_dotenv()

import re
import json
import base64
import openpyxl
import time
from openai import OpenAI
from pdf2image import convert_from_path
from PIL import Image
from io import BytesIO

# ── 설정 ──────────────────────────────────────────
DATA_DIR = os.environ["DATA_DIR"]

OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
BASE_FOLDER    = DATA_DIR
OUTPUT_EXCEL   = os.path.join(DATA_DIR, "유형분류.xlsx")
QUESTION_FOLDER = os.path.join(DATA_DIR, "문제모음")

LEVELS = ["고졸", "중졸", "초졸"]
YEARS  = ["2021", "2022", "2023", "2024", "2025", "2026"]
EXAMS_PER_LEVEL = 11 

client = OpenAI(api_key=OPENAI_API_KEY)

# ── 페이지 위치 계산 ──
def get_page_count(subject: str, exam_type: str) -> int:
    if subject == "국어" and exam_type in ("고졸", "중졸"):
        return 4
    return 2

def get_start_page_index(subject: str, exam_type: str, year: str, exam_round: str) -> int:
    if exam_type not in LEVELS:
        raise ValueError(f"알 수 없는 급수: {exam_type}")

    page_count = get_page_count(subject, exam_type)

    offset = 0
    for lvl in LEVELS:
        if lvl == exam_type:
            break
        offset += get_page_count(subject, lvl) * EXAMS_PER_LEVEL

    year_idx = YEARS.index(str(year))
    if year_idx < 5:
        within_offset = (year_idx * 2 + (int(exam_round) - 1)) * page_count
    else:
        within_offset = (5 * 2) * page_count

    return offset + within_offset

# ── 이미지 → base64 변환 ──────────────────────────
def image_to_base64(img: Image.Image) -> str:
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")

# ── GPT: 페이지 전체 보기 추출 ───────────────────
def extract_choices_from_page(page_images: list, max_retry=3) -> dict:
    prompt = """이 시험지 이미지에서 모든 문제의 보기(①②③④)를 추출하세요.

규칙:
- 문제번호는 시험지에 적힌 숫자 그대로 사용
- 반드시 이 형식으로 출력 (JSON):
{"1": "① 내용 || ② 내용 || ③ 내용 || ④ 내용", "2": "① 내용 || ..."}
- 보기가 이미지(도형/그래프/그림), 입력이 불가능한 수식, 텍스트와 이미지 혼합인 경우 빈 문자열 ""
- 보기 항목은 반드시 ①②③④ 기호를 포함할 것
- 문제번호와 보기를 정확히 매칭할 것
- 지어서 쓰지 말고, 시험지에 있는 그대로 추출할 것
- 다른 텍스트 없이 JSON만 출력"""

    content = [{"type": "text", "text": prompt}]
    for img in page_images:
        b64 = image_to_base64(img)
        content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}})

    for attempt in range(max_retry):
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": content}],
                max_tokens=2000
            )
            text = response.choices[0].message.content.strip()
            text = re.sub(r"```json|```", "", text).strip()
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if match:
                text = match.group(0)
            return json.loads(text)
        except Exception as e:
            print(f"  [재시도 {attempt+1}/{max_retry}] {e}")
            time.sleep(10)

    return {}

# ── 엑셀 로드 ──────────────────────────────────────
wb = openpyxl.load_workbook(OUTPUT_EXCEL)
ws = wb.active

headers = [cell.value for cell in ws[1]]
choices_col    = headers.index("choices") + 1
id_col         = headers.index("id") + 1
subject_col    = headers.index("subject") + 1
q_num_col      = headers.index("question_number") + 1
exam_type_col  = headers.index("exam_type") + 1
year_col       = headers.index("year") + 1
exam_round_col = headers.index("exam_round") + 1

# ── 행 데이터를 시험(급수/연도/회차/과목)별로 그룹화 ──
from collections import defaultdict
exam_rows = defaultdict(list)

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    exam_type  = row[exam_type_col - 1].value
    year       = row[year_col - 1].value
    exam_round = row[exam_round_col - 1].value
    subject    = row[subject_col - 1].value
    q_num      = row[q_num_col - 1].value

    if not exam_type:
        continue

    exam_key = (exam_type, str(year), str(exam_round), subject)
    exam_rows[exam_key].append((row_idx, q_num))

# ── 메인: 시험별로 보기 재추출 ───────────────────
total_exams = len(exam_rows)
print(f"총 {total_exams}개 시험 처리 시작\n")

# 과목별로 전체.pdf를 한 번씩만 열어서 재사용
pdf_cache = {}

for exam_idx, (exam_key, rows) in enumerate(sorted(exam_rows.items()), start=1):
    exam_type, year, exam_round, subject = exam_key

    print(f"[{exam_idx}/{total_exams}] {exam_type}_{year}_{exam_round}차_{subject}")

    if subject not in pdf_cache:
        pdf_path = os.path.join(QUESTION_FOLDER, f"{subject}전체.pdf")
        if not os.path.exists(pdf_path):
            print(f"  [PDF없음] {pdf_path}")
            pdf_cache[subject] = None
        else:
            pdf_cache[subject] = convert_from_path(pdf_path)

    all_pages = pdf_cache[subject]
    if all_pages is None:
        continue

    start_idx = get_start_page_index(subject, exam_type, year, exam_round)
    page_count = get_page_count(subject, exam_type)
    target_pages = all_pages[start_idx: start_idx + page_count]

    if not target_pages:
        print(f"  [페이지없음] idx={start_idx}")
        continue

    choices_map = extract_choices_from_page(target_pages)
    print(f"  추출된 보기: {len(choices_map)}개")

    # 엑셀 업데이트
    for row_idx, q_num in rows:
        choices = choices_map.get(str(q_num), "")
        ws.cell(row=row_idx, column=choices_col).value = choices

    time.sleep(3)

    # 10개 시험마다 저장
    if exam_idx % 10 == 0:
        wb.save(OUTPUT_EXCEL)
        print(f"  중간저장 완료")

wb.save(OUTPUT_EXCEL)
print("\n전체 완료!")