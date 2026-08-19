import os
from dotenv import load_dotenv
load_dotenv()

import re
import json
import time
import openpyxl
from anthropic import Anthropic

# ── 설정 ──────────────────────────────────────────
DATA_DIR      = os.environ["DATA_DIR"]
CODE_DIR      = os.path.dirname(os.path.abspath(__file__))

ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
EXCEL_PATH        = os.path.join(DATA_DIR, "유형분류.xlsx")
TAXONOMY_PATH     = os.path.join(CODE_DIR, "taxonomy.md")

MODEL      = "claude-haiku-4-5-20251001"
BATCH_SIZE = 15

client = Anthropic(api_key=ANTHROPIC_API_KEY)

# ── 과목별 판단 기준(few-shot 노트) ─────────────────
SUBJECT_NOTES = {
    "과학": """
[분류 시 참고할 판단 기준]
- 화살표(→) 양쪽이 같은 물질(상태만 다름)이면 "물질의 상태 변화", 다른 물질(새 물질 생성)이면 "화학 반응의 규칙과 에너지 변화"(중등) 또는 해당 반응 카테고리(고등)
- 세포호흡 반응식(포도당+산소→이산화탄소+물+에너지)은 화학이 아니라 생물로 분류 (물질대사/광합성과 호흡 등)
- "공유결합/이온결합" 개념 자체를 묻는 문제는 화학 > 화학 결합 (원자의 구조 아님)
- "물질대사"라는 용어 자체를 정의하는 문제만 생물 > 물질대사. 광합성/호흡의 구체적 과정은 기존 카테고리(광합성과 호흡) 유지
- 생태계, 생물다양성, 환경 요인에 따른 생물의 적응/진화(자연선택 포함)는 대분류가 "지구과학"이고 중분류는 "생태계" 또는 "다양성"
- 연료전지 문제: "무슨 반응/기체가 나오는가" 물으면 산화 환원, "무슨 발전 방식인가" 물으면 재생에너지(발전방식)
- "태양 고도"나 "낮의 길이" 개념이 나오면 계절의 변화. 하루 단위 자전 현상(낮밤, 동→서로 움직이는 것처럼 보임)은 지구와 달. 다른 행성/별 비교는 태양계와 별
- 힘의 종류(중력/마찰력/탄성력 등) 자체를 묻는 문제는 여러 가지 힘. 속력/위치·운동에너지 계산은 운동과 에너지
- 위치·운동에너지만 다루면 운동과 에너지(고등은 역학적에너지), 전기/화학/열/빛 등 다른 종류 에너지 간 전환이면 에너지 전환과 보존
- 광물의 특성(조흔색, 굳기, 염산 반응)·암석·지층·화산·지진은 지권의 변화. 바닷물·염류·해류·수온약층은 수권과 해수의 순환
""",
    "영어": """
[분류 시 참고할 판단 기준]
- 지문(대화/글) 안의 사실 정보를 정확히 찾아 확인/대조하는 문제 → 독해 > 세부 내용 파악
- 빈칸에 들어갈 자연스러운 대화 표현(질문이든 대답이든)을 고르는 문제, 실생활 상황 대화 → 생활영어
- 지문 없이 그림+단어만으로 판단하는 문제(그림-낱말 매칭, 공통 철자 찾기 등) → 어휘 > 단어
- 두 단어 사이의 의미 관계(유의어/반의어/상위-하위어 등) 유추 → 어휘 > 두 단어의 관계
- 비교급/최상급 문법 형태를 채우거나 판단 → 문법 > 형용사, 부사, 비교
- 대명사·지시어가 가리키는 대상을 찾는 문제 → 독해 > 지칭 추론
- 광고문/안내문/초대장/일정표 등 실물 문서 형식 지문(대화 형식 아님) → 독해 > 실용문
""",
}

# ── Taxonomy 파싱: 대분류 -> 중분류 리스트 (쌍 유지) ─────
def parse_taxonomy(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    result = {}
    section = None

    for line in text.split("\n"):
        m = re.match(r"^## (.+)", line)
        if m:
            section = m.group(1).strip()
            result[section] = {}
            continue

        if section and line.strip().startswith("- "):
            parts = line.strip()[2:].split(" > ")
            if len(parts) >= 2:
                large, mid = parts[0].strip(), parts[1].strip()
                result[section].setdefault(large, [])
                if mid not in result[section][large]:
                    result[section][large].append(mid)
            elif parts and parts[0].strip():
                large = parts[0].strip()
                result[section].setdefault(large, [])

    return result

TAXONOMY = parse_taxonomy(TAXONOMY_PATH)

def get_section_key(exam_type: str, subject: str) -> str:
    level_map = {"초졸": "초등", "중졸": "중등", "고졸": "고등"}
    level = level_map.get(exam_type, "")
    if subject == "영어":
        for key in TAXONOMY:
            if "영어" in key:
                return key
    return f"{level}_{subject}"

def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

# ── Claude 호출: 대분류 배치 분류 (tool_choice로 강제) ──
def classify_large_batch(items, large_list, subject_note, max_retry=3):
    tool = {
        "name": "assign_large_category",
        "description": "각 문제(row_idx)에 대분류를 하나씩 배정한다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "row_idx": {"type": "integer"},
                            "category_large": {"type": "string", "enum": large_list},
                        },
                        "required": ["row_idx", "category_large"],
                    },
                }
            },
            "required": ["results"],
        },
    }

    items_text = "\n\n".join(
        f"[row_idx={it['row_idx']}]\n문제: {it['question']}\n보기: {it['choices']}"
        for it in items
    )
    prompt = f"""아래 검정고시 문제들 각각에 대해 가장 적합한 대분류를 선택하세요.

사용 가능한 대분류 목록 (이 중에서만 선택):
{large_list}
{subject_note}
문제 목록:
{items_text}

각 문제(row_idx)에 대해 category_large를 하나씩 배정하세요. 반드시 assign_large_category 도구를 호출하세요."""

    for attempt in range(max_retry):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=4000,
                tools=[tool],
                tool_choice={"type": "tool", "name": "assign_large_category"},
                messages=[{"role": "user", "content": prompt}],
            )
            for block in resp.content:
                if block.type == "tool_use":
                    data = block.input
                    return {r["row_idx"]: r["category_large"] for r in data["results"]}
        except Exception as e:
            print(f"  [대분류 재시도 {attempt + 1}/{max_retry}] {e}")
            time.sleep(8)
    return {}

# ── Claude 호출: 중분류 배치 분류 ──────────────────
def classify_mid_batch(items, mid_list, subject_note, max_retry=3):
    tool = {
        "name": "assign_mid_category",
        "description": "각 문제(row_idx)에 중분류를 하나씩 배정한다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "row_idx": {"type": "integer"},
                            "category_mid": {"type": "string", "enum": mid_list},
                        },
                        "required": ["row_idx", "category_mid"],
                    },
                }
            },
            "required": ["results"],
        },
    }

    items_text = "\n\n".join(
        f"[row_idx={it['row_idx']}]\n문제: {it['question']}\n보기: {it['choices']}"
        for it in items
    )
    prompt = f"""아래 검정고시 문제들은 이미 대분류가 확정된 상태입니다.
각 문제에 대해 그 대분류에 속하는 중분류 중 가장 적합한 것을 선택하세요.

사용 가능한 중분류 목록 (이 중에서만 선택, 반드시 이 목록 안에서):
{mid_list}
{subject_note}
문제 목록:
{items_text}

각 문제(row_idx)에 대해 category_mid를 하나씩 배정하세요. 반드시 assign_mid_category 도구를 호출하세요."""

    for attempt in range(max_retry):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=4000,
                tools=[tool],
                tool_choice={"type": "tool", "name": "assign_mid_category"},
                messages=[{"role": "user", "content": prompt}],
            )
            for block in resp.content:
                if block.type == "tool_use":
                    data = block.input
                    return {r["row_idx"]: r["category_mid"] for r in data["results"]}
        except Exception as e:
            print(f"  [중분류 재시도 {attempt + 1}/{max_retry}] {e}")
            time.sleep(8)
    return {}

# ── 메인 ──────────────────────────────────────────
def main():
    backup_path = EXCEL_PATH.replace(".xlsx", "_backup.xlsx")
    if not os.path.exists(backup_path):
        import shutil
        shutil.copy(EXCEL_PATH, backup_path)
        print(f"백업 생성: {backup_path}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["exam_type", "subject", "question", "choices", "category_large", "category_mid"]
    for r in required:
        if r not in col:
            raise ValueError(f"엑셀에 '{r}' 컬럼이 없습니다. 헤더 확인 필요: {header}")

    TARGET_SUBJECTS = None 

    groups = {}
    for row in ws.iter_rows(min_row=2):
        exam_type = row[col["exam_type"]].value
        subject = row[col["subject"]].value
        if not exam_type or not subject:
            continue
        if TARGET_SUBJECTS and subject not in TARGET_SUBJECTS:
            continue
        key = (exam_type, subject)
        groups.setdefault(key, []).append(row)

    for (exam_type, subject), rows in groups.items():
        section_key = get_section_key(exam_type, subject)
        tax = TAXONOMY.get(section_key)
        if not tax:
            print(f"[taxonomy 없음] {exam_type} {subject} (key={section_key})")
            continue

        large_list = list(tax.keys())
        subject_note = SUBJECT_NOTES.get(subject, "")
        print(f"\n=== {exam_type} {subject} ({len(rows)}문항) — 대분류 {len(large_list)}개 ===")

        items = [
            {
                "row_idx": r[0].row,
                "question": (r[col["question"]].value or "")[:800],
                "choices": r[col["choices"]].value or "",
            }
            for r in rows
        ]

        large_result = {}
        for batch in chunk(items, BATCH_SIZE):
            res = classify_large_batch(batch, large_list, subject_note)
            large_result.update(res)
            time.sleep(1)

        by_large = {}
        for it in items:
            large = large_result.get(it["row_idx"])
            if large not in tax:
                print(f"  [경고] row {it['row_idx']} 대분류 미확정/유효하지 않음: {large}")
                continue
            by_large.setdefault(large, []).append(it)

        mid_result = {}
        for large, sub_items in by_large.items():
            mid_list = tax[large]
            if not mid_list:
                print(f"  [알림] '{large}' 하위에 정의된 중분류 없음 — 스킵")
                continue
            for batch in chunk(sub_items, BATCH_SIZE):
                res = classify_mid_batch(batch, mid_list, subject_note)
                mid_result.update(res)
                time.sleep(1)

        for it in items:
            row_num = it["row_idx"]
            large = large_result.get(row_num, "")
            mid = mid_result.get(row_num, "")
            ws.cell(row=row_num, column=col["category_large"] + 1, value=large)
            ws.cell(row=row_num, column=col["category_mid"] + 1, value=mid)

        wb.save(EXCEL_PATH)
        print(f"  저장완료: {exam_type} {subject}")

    print("\n전체 완료!")

if __name__ == "__main__":
    main()